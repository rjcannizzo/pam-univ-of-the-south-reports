"""
Write reports for The University of the South. Requires two input files plus the 'Key' tab's csv file.
2021-11-08
"""
import csv
import logging
from openpyxl import load_workbook, Workbook
import pandas as pd
import re
import sqlite3
from pathlib import Path
from parsers import parse_float, parse_int, parse_string

logger = logging.getLogger(__name__)
logger.setLevel(logging.INFO)
s_handler = logging.StreamHandler()
f_handler = logging.FileHandler('logs/app.log')
f_handler.setLevel(logging.INFO)
formatter = logging.Formatter('%(asctime)s - %(levelname)s - %(name)s - %(message)s', datefmt='%Y-%m-%d %H:%M:%S')
s_handler.setFormatter(formatter)
f_handler.setFormatter(formatter)
# logger.addHandler(s_handler)
logger.addHandler(f_handler)

u_south_1 = r"D:\Python\projects\pam\2021\univ_south_report_tool\examples\Univ of South Group July 2021.xlsx"
u_south_2 = r"D:\Python\projects\pam\2021\univ_south_report_tool\examples\Univ of South Individual July 2021.xlsx"

file_data = {'Group Combined': None, 'McClurg': None, 'Pub': None, 'Stirlings': None, 'Cup Gown': None,	'St Andrews': None, 'Key': None}

def get_dict_from_database(filepath):    
    """
    Return a dictionary with item numbers for keys and categories for values.
    """
    conn = sqlite3.connect(filepath)
    cursor = conn.cursor()
    cursor.execute("SELECT item, category from category;")
    data = {str(key): value for key, value in cursor.fetchall()}
    conn.close()
    return data    

CATEGORIES = get_dict_from_database(r'D:\Python\projects\pam\2021\univ_south_report_tool\categories.db')

def get_header():
    """
    Return a list of column headers. The header should have 16 columns.
    """
    return ['Category','Item #','Pack','Size','Brand','Description','MPC Code','CW','Cs Qty','Cs Total $','Cs Avg $','Split Qty','Split Total $','Split Avg $','Weight','Total Sales $']
    

def column_is_integer(text):
    """Return True if a column contains an integer else False."""
    try:
        int(text)
    except ValueError:
        return False
    return True


def convert_to_csv(input_file, output_file, skiprows=7):
    """
    Convert a tab in an Excel file to a csv.
    :param skiprows: number of rows to skip when reading the Excel file.
    :param input_file: full path to the Excel file to convert.
    :param output_file: full path of the csv file to write.
    :return: path to csv file as a string?
    """        
    df = pd.read_excel(input_file, skiprows=skiprows, usecols="A:N,P")
    df.to_csv(output_file, index=False, encoding='utf-8')
    return output_file


def create_excel_file(report_file_path):
    """
    Create an Excel file with a tab for every entry in the file_data dictionary IF there's data for the tab.
    Also adds a 'Key' tab that's a listing of categories and their meaning.
    """
    with pd.ExcelWriter(report_file_path) as writer: 
        for site, data in file_data.items():
            if not data:
                continue
            logger.info(site)
            if site != 'Key':                     
                df = pd.DataFrame(data, columns=get_header()) 
                df.to_excel(writer, sheet_name=site, index=False)                
            else:                
                df = pd.DataFrame(data)            
                df.to_excel(writer, sheet_name=site, index=False, header=False)
               

def find_first_row_of_data(filepath):
    """
    Finds the first row of data for University of the South input Excel files.
    We return row_count + 1 because the header is 2 rows, with one row after the one with 'Item #'.
    There always seems to be a 5-row section for the Sysco image. Then comes a header of 1 or more rows.
    This function won't work for other files like those for Dairy Queen. 
    """
    row_count = 0
    wb = load_workbook(filepath, data_only=True)
    first_sheet = wb.sheetnames[0]
    sheet = wb[first_sheet]    
    rows = sheet.values
    for row in rows:        
        row_count += 1
        value = row[0]
        if value == 'Item #':
            wb.close()
            return row_count + 1


def get_category(item_number):
    """Return the category for an item number if it's in the database else 'NA'.
    Categories are returned as integers.
    """
    category = CATEGORIES.get(item_number, 'NA')
    if category != 'NA':
        category = int(category)
    return category


def parse_row(row):
	"""
	Return a list of data cleaned by the specified parsers (e.g. parse_int)		
	"""
	column_parsers = [parse_string, parse_int, parse_string, parse_string, parse_string, parse_string, parse_string,
        parse_int, parse_float, parse_float, parse_int, parse_float, parse_float, parse_float, parse_float, parse_float]
	return [func(field) for func, field in zip(column_parsers, row)]


def parse_file_2(filepath):
    """
    Read a csv file with data for the 'site' tabs. Each tab's data is added to the file_data dictionary.
    """
    location = None
    data = []
    with open(filepath, encoding='utf-8') as f:
        reader = csv.reader(f)
        for row in reader:
            if not column_is_integer(row[0]):
                site = row_is_location(row[0])
                if site and data:
                    file_data[location] = data
                    data = []
                    location = site
                if site and not data:
                    location = site
                continue
            else:                
                case_quantity = float(row[7])
                split_quantity = float(row[10])
                if any([case_quantity > 0, split_quantity > 0]):
                    row = [item for item in row if not item.startswith('Unnamed')]
                    row = parse_row(row)  
                    row.insert(0, get_category(row[0]))                                      
                    data.append(row)

        file_data[location] = data


def parse_file_1(filepath):
    """
    Parses row data from a csv file and adds it to the file_data dictionary. 
    This is for the 'Group Combined' tab (tab 1) of the report.
    """
    data = []
    with open(filepath, encoding='utf-8') as f:
        reader = csv.reader(f)
        for row in reader:
            if row[0].startswith('Count'):
                break
            else:                
                case_quantity = float(row[7])
                split_quantity = float(row[10])
                if any([case_quantity > 0, split_quantity > 0]):
                    row = [item for item in row if not item.startswith('Unnamed')]                     
                    row = parse_row(row)     
                    row.insert(0, get_category(row[0]))              
                    data.append(row)
    file_data['Group Combined'] = data


def parse_key_file(filepath):
    """
    Add the data needed for the 'Key' tab to file_data dictionary.
    """
    data = []
    with open(filepath, encoding='utf-8') as f:
        reader = csv.reader(f)
        for row in reader:                        
            data.append(row)
    file_data['Key'] = data


def row_is_location(text):
    """Return True if a row is a location header row else False."""
    sites = {"ST. ANDREWS": "St Andrews", "STIRLING'S COFFEE HOUSE": "Stirlings", "CUP & GOWN": "Cup Gown", "SOUTH MCCLURG DINING": "McClurg", "PUB": "Pub" }
    for site in sites.keys():
        mo = re.search(site, text)
        if mo:
            return sites[site]
    return False


def run_report(file_1, file_2, key_file, report_file_path):
    """
    Manager function to run all necessary functions to write the report.
    """   
    parse_file_1_openpyxl(file_1)    
    parse_file_2_openpyxl(file_2)
    parse_key_file(key_file)    
    create_excel_file(report_file_path)       


def example_usage():
    " Requires two input Excel files plus the 'Key' tab's csv file."
    file_1 = r'D:\Python\projects\pam\2021\univ_south_report_tool\report_data\october\2021-10-group.xlsx'
    file_2 = r'D:\Python\projects\pam\2021\univ_south_report_tool\report_data\october\2021-10-individual.xlsx'   
    key_file = r'D:\Python\projects\pam\2021\univ_south_report_tool\report_data\key.csv'
    report_file = r'reports/univ-south-2021-10.xlsx'
    run_report(file_1, file_2, key_file, report_file_path=report_file) 


def parse_file_1_openpyxl(filepath):
    data = []    
    rows = get_file_data_as_list(filepath)
    for row in rows:           
        if row[0].startswith('Count'):
            break
        else:                
            case_quantity = row[7]
            split_quantity = row[10]
            if any([case_quantity > 0, split_quantity > 0]):
                total_sales = row[15]
                row = list(row[:14]) 
                row.append(total_sales)
                row.insert(0, get_category(row[0]))              
                data.append(row)

    file_data['Group Combined'] = data      


def get_file_data_as_list(filepath):
    """
    Return a list of file data, skipping the header and any rows that do not have a value in column 1.
    Header rows are determined by the function find_first_row_of_data().    
    """
    rows_to_skip = find_first_row_of_data(filepath)
    wb = load_workbook(filepath, data_only=True)          
    sheet = wb[wb.sheetnames[0]]
    rows = sheet.values
    for _ in range(rows_to_skip):
        next(rows)
    rows = [row for row in rows if row[0]]  
    wb.close()   
    return rows


def parse_file_2_openpyxl(filepath):
    """
    Read a csv file with data for the 'site' tabs. Each tab's data is added to the file_data dictionary.
    """
    location = None
    data = []
    # rows_to_skip = find_first_row_of_data(filepath)
    # wb = load_workbook(filepath, data_only=True)          
    # sheet = wb[wb.sheetnames[0]]
    # rows = sheet.values    
    # for _ in range(rows_to_skip):
    #     next(rows)
    rows = get_file_data_as_list(filepath)
    logger.info(rows[:10])
    for row in rows:  
        if not column_is_integer(row[0]):
            site = row_is_location(row[0])
            if site and data:
                file_data[location] = data
                data = []
                location = site
            if site and not data:
                location = site
            continue
        else:                
            case_quantity = row[7]
            split_quantity = row[10]                
            if any([case_quantity > 0, split_quantity > 0]):
                total_sales = row[15]
                row = list(row[:14]) 
                row.append(total_sales)                    
                row.insert(0, get_category(row[0]))                                      
                data.append(row)

    file_data[location] = data

def openpy_example():
    file_1 = r'D:\Python\projects\pam\2021\univ_south_report_tool\report_data\august\group.xlsx'
    file_2 = r'D:\Python\projects\pam\2021\univ_south_report_tool\report_data\august\individual.xlsx'   
    key_file = r'D:\Python\projects\pam\2021\univ_south_report_tool\report_data\key.csv'
    report_file = r'reports/univ-south-2021-08.xlsx'    
    parse_file_1_openpyxl(file_1)    
    parse_file_2_openpyxl(file_2)
    parse_key_file(key_file)    
    create_excel_file(report_file)  


def main():
    pass


if __name__ == '__main__':
    main()