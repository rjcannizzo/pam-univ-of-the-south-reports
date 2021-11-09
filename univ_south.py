"""
Write reports for The University of the South. Requires two input files plus the 'Key' tab's csv file.
2021-11-09
"""
import csv
import logging
from os import stat
from openpyxl import load_workbook
import pandas as pd
import re
import sqlite3
from pathlib import Path
from parsers import parse_float, parse_int, parse_string

class USREPORTER:
    def __init__(self, file_1, file_2, key_file, output_file_path, category_db) -> None:
        self.file_1 = file_1
        self.file_2 = file_2
        self.key_file = key_file
        self.output_file_path = output_file_path
        self.file_data = {'Group Combined': None, 'McClurg': None, 'Pub': None, 'Stirlings': None, 'Cup Gown': None,	'St Andrews': None, 'Key': None}
        self.category_db = category_db
        self.CATEGORIES = self.get_dict_from_database(self.category_db)

    @staticmethod
    def get_dict_from_database(filepath):    
        """
        Return a dictionary with item numbers for keys and categories for values.
        """
        conn = sqlite3.connect(filepath)
        cursor = conn.cursor()
        cursor.execute("SELECT item, category from category;")
        data = {str(key): value for key, value in cursor.fetchall()}
        conn.close()
        return data 

    @staticmethod
    def get_header():
        """
        Return a list of column headers. The header should have 16 columns.
        """
        return ['Category','Item #','Pack','Size','Brand','Description','MPC Code','CW','Cs Qty','Cs Total $','Cs Avg $','Split Qty','Split Total $','Split Avg $','Weight','Total Sales $']
    
    @staticmethod
    def column_is_integer(text):
        """Return True if a column contains an integer else False."""
        try:
            int(text)
        except ValueError:
            return False
        return True

    @staticmethod
    def convert_to_csv(input_file, output_file, skiprows=7):
        """
        Convert a tab in an Excel file to a csv.
        :param skiprows: number of rows to skip when reading the Excel file.
        :param input_file: full path to the Excel file to convert.
        :param output_file: full path of the csv file to write.
        :return: path to csv file as a string?
        """        
        df = pd.read_excel(input_file, skiprows=skiprows, usecols="A:N,P")
        df.to_csv(output_file, index=False, encoding='utf-8')
        return output_file

    def create_excel_file(self, report_file_path):
        """
        Create an Excel file with a tab for every entry in the file_data dictionary IF there's data for the tab.
        Also adds a 'Key' tab that's a listing of categories and their meaning.
        """
        with pd.ExcelWriter(report_file_path) as writer: 
            for site, data in self.file_data.items():
                if not data:
                    continue
                if site != 'Key':           
                    df = pd.DataFrame(data, columns = self.get_header()) 
                    df.to_excel(writer, sheet_name=site, index=False)
                else:
                    df = pd.DataFrame(data)            
                    df.to_excel(writer, sheet_name=site, index=False, header=False)

    def get_category(self, item_number):
        """Return the category for an item number if it's in the database else 'NA'.
        Categories are returned as integers.
        """
        category = self.CATEGORIES.get(item_number, 'NA')
        if category != 'NA':
            category = int(category)
        return category

    def get_file_data_as_list(self, filepath):
        """
        Return a list of file data, skipping the header and any rows that do not have a value in column 1.
        Header rows are determined by the function find_first_row_of_data().    
        """
        rows_to_skip = self.find_first_row_of_data(filepath)
        wb = load_workbook(filepath, data_only=True)          
        sheet = wb[wb.sheetnames[0]]
        rows = sheet.values
        for _ in range(rows_to_skip):
            next(rows)
        rows = [row for row in rows if row[0]]  
        wb.close()   
        return rows

    @staticmethod
    def find_first_row_of_data(filepath):
        """
        Finds the first row of data for University of the South input Excel files.
        We return row_count + 1 because the header is 2 rows, with one row after the one with 'Item #'.
        There always seems to be a 5-row section for the Sysco image. Then comes a header of 1 or more rows.
        This function won't work for other files like those for Dairy Queen. 
        """
        row_count = 0
        wb = load_workbook(filepath, data_only=True)
        first_sheet = wb.sheetnames[0]
        sheet = wb[first_sheet]    
        rows = sheet.values
        for row in rows:        
            row_count += 1
            value = row[0]
            if value == 'Item #':
                wb.close()
                return row_count + 1
    
    @staticmethod
    def parse_row(row):
        """
        Return a list of data cleaned by the specified parsers (e.g. parse_int)		
        """
        column_parsers = [parse_string, parse_int, parse_string, parse_string, parse_string, parse_string, parse_string,
            parse_int, parse_float, parse_float, parse_int, parse_float, parse_float, parse_float, parse_float, parse_float]
        return [func(field) for func, field in zip(column_parsers, row)]

    def parse_file_1(self, filepath):
        """
        Parses row data from a csv file and adds it to the file_data dictionary. 
        This is for the 'Group Combined' tab (tab 1) of the report.
        """
        data = []    
        rows = self.get_file_data_as_list(filepath)
        for row in rows:           
            if row[0].startswith('Count'):
                break
            else:                
                case_quantity = row[7]
                split_quantity = row[10]
                if any([case_quantity > 0, split_quantity > 0]):
                    total_sales = row[15]
                    row = list(row[:14]) 
                    row.append(total_sales)
                    row.insert(0, self.get_category(row[0]))              
                    data.append(row)

        self.file_data['Group Combined'] = data      

    def parse_file_2(self, filepath):
        """
        Read a csv file with data for the 'site' tabs. Each tab's data is added to the file_data dictionary.
        """
        location = None
        data = []        
        rows = self.get_file_data_as_list(filepath)        
        for row in rows:  
            if not self.column_is_integer(row[0]):
                site = self.row_is_location(row[0])
                if site and data:
                    self.file_data[location] = data
                    data = []
                    location = site
                if site and not data:
                    location = site
                continue
            else:                
                case_quantity = row[7]
                split_quantity = row[10]                
                if any([case_quantity > 0, split_quantity > 0]):
                    total_sales = row[15]
                    row = list(row[:14]) 
                    row.append(total_sales)                    
                    row.insert(0, self.get_category(row[0]))                                      
                    data.append(row)

        self.file_data[location] = data    

    def parse_key_file(self, filepath):
        """
        Add the data needed for the 'Key' tab to file_data dictionary.
        """
        data = []
        with open(filepath, encoding='utf-8') as f:
            reader = csv.reader(f)
            for row in reader:                        
                data.append(row)
        self.file_data['Key'] = data

    @staticmethod
    def row_is_location(text):
        """Return True if a row is a location header row else False."""
        sites = {"ST. ANDREWS": "St Andrews", "STIRLING'S COFFEE HOUSE": "Stirlings", "CUP & GOWN": "Cup Gown", "SOUTH MCCLURG DINING": "McClurg", "PUB": "Pub" }
        for site in sites.keys():
            mo = re.search(site, text)
            if mo:
                return sites[site]
        return False

    def run_report(self, file_1, file_2, key_file, report_file_path):
        """
        Manager function to run all necessary functions to write the report.
        """   
        self.parse_file_1(file_1)        
        self.parse_file_2(file_2)  
        self.parse_key_file(key_file)    
        self.create_excel_file(report_file_path)       

def example_usage():
    " Requires two input Excel files plus the 'Key' tab's csv file."
    file_1 = r'D:\Python\projects\pam\2021\univ_south_report_tool\report_data\october\2021-10-group.xlsx'
    file_2 = r'D:\Python\projects\pam\2021\univ_south_report_tool\report_data\october\2021-10-individual.xlsx'   
    key_file = r'D:\Python\projects\pam\2021\univ_south_report_tool\report_data\key.csv'
    report_file = r'reports/univ-south-2021-10.xlsx'
    db = r'D:\Python\projects\pam\2021\univ_south_report_tool\categories.db'
    
    reporter = USREPORTER(file_1, file_2, key_file, report_file, category_db=db)
    reporter.run_report(file_1, file_2, key_file, report_file_path=report_file) 

def main():
    example_usage()
    
    

if __name__ == '__main__':
    main()

  
    