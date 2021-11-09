"""
I'd like to find the first row of data in Excel files.
Also, how to add a table t the data on a tab.
"""
import pandas as pd
import openpyxl

file_1 = r'D:\Python\projects\pam\2021\univ_south_report_tool\report_data\october\2021-10-group.xlsx'
file_2 = r'D:\Python\projects\pam\2021\univ_south_report_tool\report_data\october\2021-10-individual.xlsx'   
funcion_wb = r'D:\Python\openpyxl\openpyxl notes\openpyxl function list.xlsx'
example_report = r'D:\Python\projects\pam\2021\univ_south_report_tool\reports\univ-south-2021-08.xlsx'

def find_first_row_of_data():
    """
    Finds the first row of data for University of the South input Excel files.
    We return row_count + 1 because the header is 2 rows, with one row after the one with 'Item #'.
    There always seems to be a 5-row section for the Sysco image. Then comes a header of 1 or more rows.
    This function won't work for other files like those for Dairy Queen. 
    """
    row_count = 0
    wb = openpyxl.load_workbook(file_2, data_only=True)
    first_sheet = wb.sheetnames[0]
    sheet = wb[first_sheet]    
    rows = sheet.values
    for row in rows:        
        row_count += 1
        value = row[0]
        if value == 'Item #':
            return row_count + 1


def table_test(filepath):
    wb = openpyxl.load_workbook(filepath, data_only=True)
    first_sheet = wb.sheetnames[0]
    sheet = wb[first_sheet]   
    print(sheet.dimensions)

def main():
    table_test(example_report)

if __name__ == '__main__':
    main()